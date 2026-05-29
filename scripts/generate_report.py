#!/usr/bin/env python3
"""
周报助手脚本 - 读取模板和生成周报
"""

import os
import sys
import json
import glob
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(SCRIPT_DIR)

DEFAULT_TEMPLATE_PATH = os.path.join(SKILL_DIR, "templates", "赛仕科工作周报模板.xlsx")
DEFAULT_OUTPUT_DIR = "/mnt/c/Users/OLIVER_SONG.AADDS/Desktop/周报/"


def get_template_path():
    env_path = os.environ.get('WKR_TEMPLATE_PATH')
    if env_path and os.path.exists(env_path):
        return env_path
    if os.path.exists(DEFAULT_TEMPLATE_PATH):
        return DEFAULT_TEMPLATE_PATH
    return os.environ.get('WKR_TEMPLATE_PATH', DEFAULT_TEMPLATE_PATH)


def get_output_dir():
    return os.environ.get('WKR_OUTPUT_DIR', DEFAULT_OUTPUT_DIR)


def calculate_saturation(last_week, this_week, next_week):
    """
    根据工作内容自动计算饱和度，最低85%
    规则：
    - 1-2项工作：85%-90%
    - 3-4项工作：90%-95%
    - 5项以上或跨多项目：95%-100%
    """
    total_items = len(last_week) + len(this_week) + len(next_week)
    
    if total_items <= 2:
        base = 85
    elif total_items <= 4:
        base = 90
    else:
        base = 95
    
    projects = set()
    for item in last_week + this_week + next_week:
        content = item.get('content', '')
        if '绩效' in content or '考核' in content:
            projects.add('performance')
        elif '督办' in content or 'CMO' in content:
            projects.add('supervision')
        elif '积分' in content or '属地' in content:
            projects.add('local')
        elif 'E-Loading' in content or 'Career' in content:
            projects.add('e-loading')
        elif 'AI' in content or 'Firebase' in content or '智能体' in content:
            projects.add('ai')
        else:
            projects.add('other')
    
    if len(projects) >= 3:
        base += 5
    
    return min(100, max(85, base))


def set_cell_value(sheet, row, col, value):
    """设置单元格值，保留换行格式"""
    cell = sheet.cell(row=row, column=col, value=value)
    if isinstance(value, str) and '\n' in value:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    return cell


def find_section_rows(sheet):
    """动态查找各板块标题行和数据起始行"""
    sections = {}
    for row in range(1, sheet.max_row + 1):
        cell_val = sheet.cell(row=row, column=1).value
        if cell_val and isinstance(cell_val, str):
            if '上周工作完成情况' in cell_val:
                sections['last_week_header'] = row
                sections['last_week_data_start'] = row + 2
            elif '本周重要工作项' in cell_val:
                sections['this_week_header'] = row
                sections['this_week_data_start'] = row + 2
            elif '下周工作计划' in cell_val:
                sections['next_week_header'] = row
                sections['next_week_data_start'] = row + 2
            elif '工作饱和度评估' in cell_val:
                sections['saturation_header'] = row
                sections['saturation_data'] = row + 1
    return sections


def fill_section(sheet, start_row, items, columns):
    """
    填充一个板块的数据
    items: list of dict
    columns: dict mapping field_name -> column_index
    """
    for i, item in enumerate(items):
        row = start_row + i
        for field, col in columns.items():
            if field in item and item[field] is not None:
                set_cell_value(sheet, row, col, item[field])


def read_last_week_report():
    """读取最新的周报文件"""
    pattern = os.path.join(OUTPUT_DIR, "赛仕科工作周报 - *.xlsx")
    files = glob.glob(pattern)
    if not files:
        return None
    # 按修改时间排序，取最新的
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0]


def migrate_from_last_week(last_week_path):
    """
    从上周周报迁移数据
    本周工作项 -> 上周完成情况
    下周计划 -> 本周工作项
    """
    wb = load_workbook(last_week_path)
    sheet = wb['员工工作周报']
    
    migrated = {
        'last_week': [],
        'this_week': []
    }
    
    # 读取本周工作项 (第13-16行)
    for row in range(13, 17):
        content = sheet.cell(row=row, column=3).value
        if content:
            migrated['last_week'].append({
                'content': content,
                'planned_date': sheet.cell(row=row, column=5).value,
                'actual_date': sheet.cell(row=row, column=6).value,
                'status': '已完成',  # 默认已完成
                'progress': 1,
                'results': '',
                'problems': sheet.cell(row=row, column=9).value,
                'solutions': '',
                'notes': sheet.cell(row=row, column=11).value
            })
    
    # 读取下周计划 (第21-24行)
    for row in range(21, 25):
        content = sheet.cell(row=row, column=3).value
        if content:
            migrated['this_week'].append({
                'content': content,
                'priority': sheet.cell(row=row, column=4).value or '中',
                'start_date': sheet.cell(row=row, column=5).value,
                'end_date': sheet.cell(row=row, column=6).value,
                'owner': sheet.cell(row=row, column=7).value,
                'collaborators': sheet.cell(row=row, column=8).value,
                'problems': sheet.cell(row=row, column=9).value,
                'eta': sheet.cell(row=row, column=10).value,
                'notes': sheet.cell(row=row, column=11).value
            })
    
    return migrated


def generate_weekly_report(report_data):
    """
    生成周报文件
    report_data 结构:
    {
        'period': '2026/5/26-2026/5/30',
        'name': 'OLIVER SONG 宋俊佑',
        'department': 'IT信息技术部',
        'last_week': [...],
        'this_week': [...],
        'next_week': [...],
        'saturation': 0.9,
        'workload_desc': '...',
        'resources_needed': '',
        'other_notes': ''
    }
    """
    wb = load_workbook(get_template_path())
    sheet = wb['员工工作周报']
    sections = find_section_rows(sheet)
    
    # 表头 (Row 2)
    sheet['B2'] = report_data.get('period', '')
    sheet['E2'] = report_data.get('name', 'OLIVER SONG 宋俊佑')
    sheet['H2'] = report_data.get('department', 'IT信息技术部')
    
    # 上周完成情况列映射
    # A:序号, B:工作内容, C:预计完成时间, D:实际完成时间, E:完成状态, F:完成进度(公式), G:成果说明, H:遇到的问题, I:解决方案, J:备注
    last_week_cols = {
        'content': 2, 'planned_date': 3, 'actual_date': 4,
        'status': 5, 'results': 7,
        'problems': 8, 'solutions': 9, 'notes': 10
    }
    fill_section(sheet, sections['last_week_data_start'], report_data.get('last_week', []), last_week_cols)
    
    # 本周工作项列映射
    # A:序号, B:工作内容, C:优先级, D:计划开始时间, E:计划完成时间, F:责任人, G:协作人, H:遇到的问题, I:预计完成时间, J:备注
    this_week_cols = {
        'content': 2, 'priority': 3, 'start_date': 4, 'end_date': 5,
        'owner': 6, 'collaborators': 7, 'problems': 8, 'eta': 9, 'notes': 10
    }
    fill_section(sheet, sections['this_week_data_start'], report_data.get('this_week', []), this_week_cols)
    
    # 下周计划列映射
    fill_section(sheet, sections['next_week_data_start'], report_data.get('next_week', []), this_week_cols)
    
    sat_row = sections['saturation_data']
    saturation = report_data.get('saturation', '')
    
    if not saturation:
        auto_saturation = calculate_saturation(
            report_data.get('last_week', []),
            report_data.get('this_week', []),
            report_data.get('next_week', [])
        )
        saturation = f"{auto_saturation}%"
    elif not str(saturation).endswith('%'):
        saturation = f"{saturation}%"
    
    sheet.cell(row=sat_row, column=2, value=saturation)
    sheet.cell(row=sat_row, column=5, value=report_data.get('workload_desc', ''))
    sheet.cell(row=sat_row + 1, column=2, value=report_data.get('resources_needed', ''))
    sheet.cell(row=sat_row + 1, column=5, value=report_data.get('other_notes', ''))
    
    period = report_data.get('period', '')
    date_suffix = datetime.now().strftime('%Y-%m-%d')
    if '-' in period:
        parts = period.split('-')
        if len(parts) == 2:
            start_date = parts[0].strip().replace('/', '-')
            end_date = parts[1].strip().replace('/', '-')
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                date_suffix = f"{start_dt.strftime('%Y-%m-%d')}至{end_dt.strftime('%Y-%m-%d')}"
            except ValueError:
                date_suffix = f"{start_date}至{end_date}"
        else:
            date_suffix = datetime.now().strftime('%Y-%m-%d')
    
    filename = f"赛仕科工作周报 - {report_data.get('name', 'OLIVER SONG 宋俊佑')} {date_suffix}.xlsx"
    output_path = os.path.join(get_output_dir(), filename)
    wb.save(output_path)
    return output_path


if __name__ == '__main__':
    # 测试用法
    if len(sys.argv) > 1 and sys.argv[1] == '--test':
        test_data = {
            'period': '2026/5/26-2026/5/30',
            'name': 'OLIVER SONG 宋俊佑',
            'department': 'IT信息技术部',
            'last_week': [
                {
                    'content': '香港绩效系统二期需求分析',
                    'planned_date': '2026-05-26',
                    'actual_date': '2026-05-26',
                    'status': '已完成',
                    'progress': 1,
                    'results': '完成需求文档撰写和评审',
                    'problems': '暂无'
                }
            ],
            'this_week': [
                {
                    'content': '属地积分管理系统开发',
                    'priority': '高',
                    'start_date': '2026-05-18',
                    'end_date': '2026-05-21',
                    'owner': 'OLIVER SONG 宋俊佑',
                    'collaborators': '',
                    'problems': '一票否决功能需求变更',
                    'eta': '2026-05-21'
                }
            ],
            'next_week': [
                {
                    'content': 'E-Loading装载系统需求分析',
                    'priority': '高',
                    'start_date': '2026-05-20',
                    'end_date': '2026-06-15',
                    'owner': 'OLIVER SONG 宋俊佑',
                    'collaborators': 'BILLY CHEN 陈广全',
                    'problems': '暂无',
                    'eta': '2026-06-15'
                }
            ],
            'saturation': 0.9,
            'workload_desc': '本周双线并行推进，进度符合预期',
            'resources_needed': '',
            'other_notes': ''
        }
        path = generate_weekly_report(test_data)
        print(f"Generated: {path}")
